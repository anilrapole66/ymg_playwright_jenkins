from dateutil.relativedelta import relativedelta
from django.shortcuts import render, get_object_or_404, redirect
from .models import Employee, MainClient, EndClient, MigrantType, LeaveType, MedicalCertificate, PublicHoliday, \
    LeaveAdjustment, ClaimAttachment, Claim, LeaveImport
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required,user_passes_test
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.urls import reverse
import random
import string
from openpyxl import Workbook

from django.shortcuts import render, redirect
from .forms import EmployeeForm, MainClientForm, EndClientForm, MigrantTypeForm, ProfileUpdateForm, LeaveAdjustmentForm, \
    ClaimForm, ClaimAttachmentForm, PublicHolidayForm
from django.forms import inlineformset_factory

from django.shortcuts import render, redirect
from .models import MainClient, EndClient, MigrantType
from .forms import MainClientForm, EndClientForm, MigrantTypeForm
from django.contrib.auth.models import User
import datetime
from employees.business_logic.leave_service import (
    update_leave_cycle,
    get_or_create_active_cycle,
    iter_working_dates,
    effective_leave_days,
)
from employees.business_logic.aws_service import send_password_reset_email
import openpyxl
from django.shortcuts import render, redirect
from .models import Employee, MainClient, EndClient, MigrantType, Manager
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str

import openpyxl

def parse_excel_date(value):
    """Safely parse date values from Excel in multiple formats."""
    if not value:
        return None

    # 1️⃣ If Excel gives a real date object (already parsed)
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value if isinstance(value, datetime.date) else value.date()

    # 2️⃣ Try common string patterns
    value = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%Y", "%d/%b/%Y", "%Y-%b-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    # 3️⃣ Fallback: try Excel serial numbers
    try:
        return openpyxl.utils.datetime.from_excel(value)
    except Exception:
        pass

    print(f"⚠️ Warning: Could not parse date '{value}'")
    return None


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def configurables_view(request):

    main_clients = MainClient.objects.all()
    end_clients = EndClient.objects.all()
    migrant_types = MigrantType.objects.all()
    rolesows = RoleSow.objects.all()

    if request.method == "POST":

        main_id = request.POST.get("main_account")
        end_id = request.POST.get("end_client")

        # Must pick at least one filter
        if not main_id and not end_id:
            messages.error(request, "Please select Main Account or End Client.", extra_tags="bulk")
            return redirect("/configurables/?tab=bulk")

        # Build employee queryset
        employees = Employee.objects.all()

        if main_id:
            employees = employees.filter(main_account_id=main_id)

        if end_id:
            employees = employees.filter(end_client_id=end_id)

        if not employees.exists():
            messages.error(request, "No employees found for the selected filters.", extra_tags="bulk")
            return redirect("/configurables/?tab=bulk")

        updates = {}

        # Boolean settings
        if request.POST.get("ot_allowed"):
            updates["ot_allowed"] = (request.POST.get("ot_allowed_value") == "true")

        if request.POST.get("ph_work_allowed"):
            updates["ph_work_allowed"] = (request.POST.get("ph_work_allowed_value") == "true")

        # Numeric fields with validation
        numeric_fields = [
            ("shift_hours_day", "shift_hours_day_value", "Shift Hours (Day)"),
            ("shift_days_week", "shift_days_week_value", "Shift Days (Week)"),
            ("annual_ent", "annual_ent_value", "Annual Leave Entitlement"),
            ("medical_ent", "medical_ent_value", "Medical Leave Entitlement")
        ]

        for checkbox, field, label in numeric_fields:
            if request.POST.get(checkbox):
                value = request.POST.get(field)
                if not value:
                    messages.error(request, f"{label} cannot be empty.", extra_tags="bulk")
                    return redirect("/configurables/?tab=bulk")
                updates[
                    "yearly_sick_leave_entitlement" if checkbox == "medical_ent"
                    else "yearly_annual_leave_entitlement" if checkbox == "annual_ent"
                    else field.replace("_value", "")
                ] = value

        # Location
        if request.POST.get("location"):
            updates["location"] = request.POST.get("location_value")

        # No fields selected
        if not updates:
            messages.error(request, "Please select at least one field to update.", extra_tags="bulk")
            return redirect("/configurables/?tab=bulk")

        # Perform update
        count = employees.update(**updates)
        messages.success(request, f"Successfully updated {count} employees.", extra_tags="bulk")

        return redirect("/configurables/?tab=bulk")

    # GET request
    selected_tab = request.GET.get("tab", "main")

    return render(request, 'employees/configurables.html', {
        'main_clients': main_clients,
        'end_clients': end_clients,
        'migrant_types': migrant_types,
        'rolesows': rolesows,
        'selected_tab': selected_tab
    })


def generate_strong_password(length=12):
    """Generate a strong random password with upper, lower, digits, and symbols."""
    chars = string.ascii_letters + string.digits + "!@#$%^&*()-_=+"
    while True:
        pwd = ''.join(random.choice(chars) for _ in range(length))
        # ensure at least one lowercase, one uppercase, one digit, one special char
        if (pwd[0] != '=' and any(c.islower() for c in pwd)
                and any(c.isupper() for c in pwd)
                and any(c.isdigit() for c in pwd)
                and any(c in "!@#$%^&*()-_=+" for c in pwd)):
            return pwd

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_rolesow(request):
    """
    View to add a new RoleSow
    """
    if request.method == 'POST':
        form = RoleSowForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('configurables')  # Redirect to the list of roles
    else:
        form = RoleSowForm()

    return render(request, 'employees/add_rolesow.html', {'form': form})

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_main_client(request):
    form = MainClientForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('configurables')  # Redirect back to configurables page
    return render(request, 'employees/add_main_client.html', {'form': form})

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_end_client(request):
    form = EndClientForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('configurables')  # Redirect to configurables page after saving

    # Fetch all MainClient instances
    main_clients = MainClient.objects.all()

    return render(request, 'employees/add_end_client.html', {
        'form': form,
        'main_clients': main_clients,  # Pass MainClient data to the template
    })

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_migrant_type(request):
    form = MigrantTypeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('configurables')  # Redirect back to configurables page
    return render(request, 'employees/add_migrant_type.html', {'form': form})


from django.contrib import messages  # Add for success messages

from openpyxl import Workbook
from django.http import HttpResponse

def upload_employees(request):
    if request.method == 'POST':
        excel_file = request.FILES['employee_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        uploaded_employees = []
        new_credentials = []  # collect credentials for HR Excel export
        duplicate_count = 0   # track duplicates by email

        raw_headers = [cell.value for cell in sheet[1]]

        headers = [h for h in raw_headers if h is not None]
        header_len = len(headers)

        print("CLEAN HEADERS:", header_len, headers)

        VALID_COUNTRIES = {c[0] for c in Employee.country_choices}

        def parse_bool(val):
            if isinstance(val, bool):
                return val
            if val is None:
                return False
            if isinstance(val, (int, float)):
                return val != 0
            s = str(val).strip().lower()
            return s in ("yes", "true", "1", "y")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = row[:header_len]

            data = dict(zip(headers, row))

            if not data.get("Full Name"):
                break

            raw_location = data.get("Country")

            location = raw_location.strip() if isinstance(raw_location, str) else raw_location
            if not location or location not in VALID_COUNTRIES:
                location = Employee.country_choices[0][0]

            employee_serial_id = data.get("Employee Serial ID") or data.get("Employee ID")
            full_name = data.get("Full Name")
            email = data.get("Email")
            phone = data.get("Phone")
            main_account_name = data.get("Main Account")
            end_client_name = data.get("End Client")
            account_manager_email = data.get("Account Manager Email")
            pass_type_val = data.get("Pass Type")
            role = data.get("Role")
            role_sow_name = data.get("Role SOW")
            doj = data.get("Date of Joining")
            is_pmo = data.get("Is PMO")
            ot_allowed = data.get("OT Allowed")
            shift_hours_day = data.get("Shift Hours")
            shift_days_week = data.get("Shift Days")
            annual_leave = data.get("Annual Leave")
            sick_leave = data.get("Sick Leave")
            ph_allowed = data.get("PH Allowed")

            if not (full_name and email):
                continue

            # ✅ Check for duplicate email before proceeding
            if Employee.objects.filter(email__iexact=email.strip()).exists():
                duplicate_count += 1
                continue  # skip duplicates safely

            # --- Foreign Keys ---
            main_account = None
            if main_account_name:
                main_account, _ = MainClient.objects.get_or_create(name=main_account_name)

            end_client = None
            if end_client_name and main_account:
                end_client, _ = EndClient.objects.get_or_create(
                    name=end_client_name,
                    main_client=main_account
                )

            pass_type = None
            if pass_type_val:
                pass_type, _ = MigrantType.objects.get_or_create(migrant_name=pass_type_val)

            role_sow = None
            if main_account:
                if role_sow_name:
                    prefixed_role_name = f"{main_account.name.strip()}-{str(role_sow_name).strip()}"
                    role_sow = RoleSow.objects.filter(
                        name__iexact=prefixed_role_name,
                        main_client=main_account,
                    ).order_by("id").first()
                    if not role_sow:
                        role_sow = RoleSow.objects.create(
                            name=prefixed_role_name,
                            description=f"Auto-created via Excel upload for {main_account.name}",
                            main_client=main_account,
                        )
                else:
                    role_sow = RoleSow.objects.filter(
                        main_client=main_account
                    ).order_by("id").first()
                    if not role_sow:
                        role_sow = RoleSow.objects.create(
                            name=main_account.name,
                            description=f"Auto-created via Excel upload for {main_account.name}",
                            main_client=main_account,
                        )

            manager = None
            if account_manager_email:
                manager = Manager.objects.filter(email=account_manager_email).first()

            # --- Create Employee ---
            employee = Employee.objects.create(
                employee_serial_id=employee_serial_id,
                full_name=full_name,
                email=email,
                phone=phone,
                main_account=main_account,
                end_client=end_client,
                client_account_manager=manager,
                pass_type=pass_type,
                work_role=role,
                role_sow=role_sow,
                date_of_joining=parse_excel_date(doj),
                is_active=True,
                is_pmo=parse_bool(is_pmo),
                ot_allowed=parse_bool(ot_allowed),
                shift_hours_day=shift_hours_day or 8,
                shift_days_week=shift_days_week or 5,
                yearly_annual_leave_entitlement=int(annual_leave or 14),
                yearly_sick_leave_entitlement=int(sick_leave or 14),
                location=location,
                ph_allowed=parse_bool(ph_allowed),
            )

            employee_from_db = Employee.objects.get(id=employee.id)
            print("DB VALUE IMMEDIATELY AFTER CREATE:", employee_from_db.location)

            # --- Create linked User if missing ---
            if not User.objects.filter(username=email).exists():
                random_password = generate_strong_password()
                user = User.objects.create_user(username=email, password=random_password)
                employee.user = user
                employee.location = employee.location
                employee.save()

                new_credentials.append({
                    "email": email,
                    "password": random_password
                })

            uploaded_employees.append(employee)

        # ✅ Export new credentials Excel
        if new_credentials:
            wb_out = Workbook()
            ws = wb_out.active
            ws.title = "Login Credentials"
            ws.append(["Email", "Generated Password"])

            for cred in new_credentials:
                ws.append([cred["email"], cred["password"]])

            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Employee_Credentials_{timestamp}.xlsx"

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb_out.save(response)
            return response

        # ✅ Add detailed message summary
        messages.success(
            request,
            f"{len(uploaded_employees)} new employee(s) uploaded successfully! "
            f"{duplicate_count} duplicate(s) skipped."
        )

        return render(request, 'employees/upload_employees.html', {
            'uploaded_employees': uploaded_employees
        })

    return render(request, 'employees/upload_employees.html')




@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_employee(request):
    # Initialize the employee form
    employee_form = EmployeeForm()

    if request.method == 'POST':
        employee_form = EmployeeForm(request.POST)
        if employee_form.is_valid():
            new_employee = employee_form.save(commit = False)
            if not new_employee.role_sow and new_employee.main_account:
                client_role = RoleSow.objects.filter(
                    main_client=new_employee.main_account
                ).order_by("id").first()
                if not client_role:
                    client_role = RoleSow.objects.create(
                        name=new_employee.main_account.name,
                        main_client=new_employee.main_account,
                    )
                new_employee.role_sow = client_role
            new_employee.is_active = True
            # Create a user with the same username and password as the email
            username = new_employee.email  # Use the email as the username for the User
            password = new_employee.email  # Temporarily set the password as the same email
            user = User.objects.create_user(username=username, password=password)
            new_employee.user = user
            new_employee.save()

            return redirect('employee_list')  # Redirect to employee list page if form is valid

    # Pass the dropdown data to the template
    main_clients = MainClient.objects.all()
    end_clients = EndClient.objects.all()
    migrant_types = MigrantType.objects.all()
    rolesows = RoleSow.objects.all()
    managers = Manager.objects.filter(is_active=True)

    return render(request, 'employees/add_employee.html', {
        'employee_form': employee_form,
        'main_clients': main_clients,
        'end_clients': end_clients,
        'migrant_types': migrant_types,
        'rolesows' : rolesows,
        'managers': managers,
    })
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@login_required
def employee_list(request):
    query = request.GET.get('q', '')
    main_account_id = request.GET.get('main_account', '')
    end_client_id = request.GET.get('end_client', '')

    employees = Employee.objects.all().order_by('-is_active', 'full_name')

    # Apply name search if 'q' is provided
    if query:
        employees = employees.filter(full_name__icontains=query)

    # Filter by Main Account if provided
    if main_account_id.isdigit():  # Ensure it is a valid integer
        main_account_id = int(main_account_id)
        employees = employees.filter(main_account_id=main_account_id)
    else:
        main_account_id = None

    # Filter by End Client if provided
    if end_client_id.isdigit():  # Ensure it is a valid integer
        end_client_id = int(end_client_id)
        employees = employees.filter(end_client_id=end_client_id)
    else:
        end_client_id = None

    # Get all MainClient and EndClient objects for dropdowns
    main_clients = MainClient.objects.all()
    end_clients = EndClient.objects.all()

    return render(request, 'employees/employee_list.html', {
        'employees': employees,
        'query': query,
        'main_clients': main_clients,
        'end_clients': end_clients,
        'selected_main_account': main_account_id,
        'selected_end_client': end_client_id,
    })

@login_required
@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    # ============================================
    # 1️⃣ ADMIN ACCESS → allowed for all
    # ============================================
    if request.user.is_staff or request.user.is_superuser:
        return render(request, "employees/employee_detail_dashboard.html", {
            "employee": employee,
            "is_admin": True,
            "is_manager": False,
            "is_employee_self": False,
            "parent_template": "employees/dashboard.html",
        })

    # ============================================
    # 2️⃣ MANAGER ACCESS → allowed only for their team
    # ============================================
    if hasattr(request.user, "manager_profile"):
        manager = request.user.manager_profile
        if employee.client_account_manager == manager:
            return render(request, "employees/employee_detail_dashboard.html", {
                "employee": employee,
                "is_admin": False,
                "is_manager": True,
                "is_employee_self": False,
                "parent_template": "employees/manager_dashboard.html",
            })
        return HttpResponseForbidden("Not allowed")


    # ============================================
    # 4️⃣ Default deny
    # ============================================
    return HttpResponseForbidden("Not allowed")


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    old_active_status = employee.is_active
    if request.method == "POST":
        employee.employee_serial_id = request.POST.get('employee_serial_id')
        employee.full_name = request.POST.get('full_name')
        employee.date_of_birth = request.POST.get('date_of_birth') or None
        employee.email = request.POST.get('email')
        employee.phone = request.POST.get('phone')
        employee.main_account_id = request.POST.get('main_account') or None
        if 'end_client' in request.POST:
            employee.end_client_id = request.POST.get('end_client') or None
        employee.is_active = request.POST.get('is_active') == 'True'
        if 'pass_type' in request.POST:
            employee.pass_type_id = request.POST.get('pass_type') or None
        if 'role_sow' in request.POST:
            employee.role_sow_id = request.POST.get('role_sow') or None
        employee.work_role = request.POST.get('work_role')
        employee.date_of_joining = request.POST.get('date_of_joining') or None
        employee.sow_start_date = request.POST.get('sow_start_date') or None
        employee.sow_end_date = request.POST.get('sow_end_date') or None

        if 'is_pmo' in request.POST:
            employee.is_pmo = request.POST.get('is_pmo') == 'on'
        employee.ot_allowed = request.POST.get('ot_allowed') == 'on'

        employee.shift_hours_day = request.POST.get('shift_hours_day') or 8
        employee.shift_days_week = request.POST.get('shift_days_week') or 5
        employee.yearly_annual_leave_entitlement = request.POST.get('yearly_annual_leave_entitlement') or 5
        employee.yearly_sick_leave_entitlement = request.POST.get('yearly_sick_leave_entitlement') or 5
        employee.location = request.POST.get('location')
        employee.ph_allowed = request.POST.get('ph_allowed') == 'on'
        employee.save()

        if employee.user and old_active_status != employee.is_active:
            employee.user.is_active = employee.is_active
            employee.user.save()
        return redirect("employee_detail", pk=employee.pk)

    context = {
        "employee": employee,
        "main_clients": MainClient.objects.all(),
        "end_clients": EndClient.objects.all(),
        "migrant_types": MigrantType.objects.all(),
        "rolesows": RoleSow.objects.all(),
    "country_choices": PublicHoliday.COUNTRY_CHOICES,
    }
    return render(request, "employees/employee_detail.html", context)


@login_required
def user_redirect(request):
    """
    Redirect users to their respective homepages.
    """
    if request.user.is_staff:  # Staff users should see the existing employee_list page
        return redirect('employee_list')  # Reuse existing admin view

        # 2️⃣ Manager users → Manager Dashboard
    if hasattr(request.user, "manager_profile"):
        return redirect('manager_employee_list')
    return redirect('employee_home')  # Redirect to the employee-specific homepage



@login_required
def edit_profile(request):
    # Get the currently logged-in employee's profile
    employee = Employee.objects.get(email=request.user.username)
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Your profile has been updated successfully!")
            return redirect('employee_home')  # Redirect back to the employee dashboard
    else:
        form = ProfileUpdateForm(instance=employee)  # Pre-fill the form with existing data

    return render(request, 'employees/edit_profile.html', {
        'form': form,
    })


from django.shortcuts import render, get_object_or_404, redirect
from .models import RoleSow, TaskType, Task
from .forms import RoleSowForm, TaskForm

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def configurables_roles(request):
    """
    List all SOW (Roles) and link to their details.
    """
    roles = RoleSow.objects.all()
    return render(request, 'employees/configurables_roles.html', {'roles': roles})

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def configurables_role_details(request, rolesow_id):
    """
    View and manage task types and tasks for a specific RoleSow.
    """
    rolesow = get_object_or_404(RoleSow, id=rolesow_id)
    task_types = rolesow.task_types.all()  # All types related to this RoleSow
    tasks = rolesow.tasks.filter()  # Only base tasks related to this RoleSow

    return render(request, 'employees/configurables_role_details.html', {
        'rolesow': rolesow,
        'task_types': task_types,
        'tasks': tasks,
    })


from .forms import TaskTypeForm, TaskForm

def add_task_type(request, rolesow_id):
    """
    View to add a new Task Type for a specific RoleSow.
    """
    rolesow = get_object_or_404(RoleSow, id=rolesow_id)

    if request.method == 'POST':
        form = TaskTypeForm(request.POST)
        if form.is_valid():
            task_type = form.save(commit=False)
            task_type.role_sow = rolesow  # Assign RoleSow to the Task Type
            task_type.save()
            return redirect('configurables_role_details', rolesow_id=rolesow.id)
    else:
        form = TaskTypeForm()

    return render(request, 'employees/add_task_type.html', {'form': form, 'rolesow': rolesow})

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_task(request, rolesow_id):
    """
    View to add a new Task for a specific RoleSow.
    """
    rolesow = get_object_or_404(RoleSow, id=rolesow_id)

    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.role_sow = rolesow  # Assign RoleSow to the Task
            task.is_base = True  # Set as base task initially
            task.save()
            return redirect('configurables_role_details', rolesow_id=rolesow.id)
    else:
        form = TaskForm(rolesow=rolesow)

    return render(request, 'employees/add_task.html', {'form': form, 'rolesow': rolesow})


from django.contrib.auth.decorators import login_required

@login_required
def saved_timesheets(request):
    employee = get_object_or_404(Employee, email=request.user.username)

    # Fetch all submissions for this employee (ordered newest first)
    submissions = (
        TimesheetSubmission.objects
        .filter(employee=employee)
        .select_related('period')
        .order_by('-period__start_date')
    )

    return render(request, 'employees/saved_timesheets.html', {
        'submissions': submissions,
    })

@login_required
def filter_tasks(request):
    task_type_id = request.GET.get('task_type_id')
    role_sow_id = request.GET.get('role_sow_id')

    if not task_type_id or not role_sow_id:
        return JsonResponse({'error': 'Task Type ID and Role SOW ID are required'}, status=400)

    # Filter tasks by both TaskType and RoleSow
    tasks = Task.objects.filter(type_id=task_type_id, role_sow_id=role_sow_id).values('id', 'name', 'description')
    return JsonResponse({'tasks': list(tasks)})

import json
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.safestring import mark_safe
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import (
    Employee, TaskType, Task, LeaveType,
    TimesheetEntry, TimesheetPeriod,
    TimesheetSubmission, TimesheetApprovalLog
)
from django.db.models import Q
from django.db import transaction
@login_required
def submit_timesheet(request):
    """Handles both displaying and saving/submitting the employee’s timesheet."""
    employee_id = request.GET.get("employee_id")
    if employee_id:
        # Someone is trying to view another employee's timesheet
        target_employee = get_object_or_404(Employee, id=employee_id)

        # --- Admin allowed for all ---
        if request.user.is_staff or request.user.is_superuser:
            current_employee = target_employee

        # --- Manager allowed ONLY if employee is under them ---
        elif hasattr(request.user, "manager_profile"):
            manager = request.user.manager_profile
            if target_employee.client_account_manager == manager:
                current_employee = target_employee
            else:
                return HttpResponseForbidden("You are not allowed to view this employee's timesheet.")

        # --- Employees cannot view others ---
        else:
            return HttpResponseForbidden("You are not allowed to view this employee's timesheet.")

    else:
        # ✅ Default: logged-in employee viewing their own sheet
        current_employee = get_object_or_404(Employee, email=request.user.username)

    role_sow = current_employee.role_sow



    # ---------------------------
    # GET → Load data for selected month
    # ---------------------------
    if request.method == "GET":
        year = request.GET.get("year")
        month = request.GET.get("month")

        employee_country = current_employee.location or "Singapore"
        if year and month:
            current_month = f"{year}-{month.zfill(2)}"
        else:
            current_month = datetime.date.today().strftime("%Y-%m")

        view_flag = request.GET.get("view")  # will be "1" when coming from the View button
        read_only = True if view_flag in ("1", "true", "True") else False

        first_day = datetime.datetime.strptime(f"{current_month}-01", "%Y-%m-%d").date()
        last_day = (first_day + relativedelta(months=1)) - datetime.timedelta(days=1)

        selected_month_date = first_day
        holidays = PublicHoliday.objects.filter(
            country=employee_country,
            holiday_date__month=selected_month_date.month,
            holiday_date__year=selected_month_date.year,
            is_active=True
        )
        holiday_map = {h.holiday_date.strftime("%Y-%m-%d"): h.holiday_name for h in holidays}

        submission = TimesheetSubmission.objects.filter(
            employee=current_employee,
            period__start_date=first_day
        ).first()
        submission_status_text = submission.status if submission else ""
        # If fully approved → always read only
        if submission and submission.status == "APPROVED":
            read_only = True
        # Fetch reference data
        task_types = TaskType.objects.filter(role_sow=role_sow)
        tasks = Task.objects.filter(role_sow=role_sow, is_active=True)
        leave_types = LeaveType.objects.all()

        # Fetch saved timesheet entries
        raw_entries = TimesheetEntry.objects.filter(
            employee=current_employee,
            date__range=(first_day, last_day)
        ).values(
            "date", "job_type", "task_type_id", "task_id",
            "work_hours", "ot_hours", "break_hours",
            "leave_type_id", "description", "is_ot"
        )
        approved_leaves = LeaveApplication.objects.filter(
            employee=current_employee,
            status="APPROVED",
            start_date__lte=last_day,
            end_date__gte=first_day,
        )
        leave_date_map = {}  # date -> leave_type_id

        for leave in approved_leaves:
            current = max(leave.start_date, first_day)
            end = min(leave.end_date, last_day)

            for work_date in iter_working_dates(current, end):
                leave_date_map[work_date.strftime("%Y-%m-%d")] = leave.leave_type_id
        # Safely convert Decimals to floats for JSON
        timesheet_entries = []
        for e in raw_entries:
            e["date"] = e["date"].strftime("%Y-%m-%d")
            e["job_type"] = e["job_type"].title()
            for key in ["work_hours", "ot_hours", "break_hours"]:
                val = e.get(key)
                e[key] = float(val) if isinstance(val, Decimal) else (val or 0)
            timesheet_entries.append(e)

        existing_dates = {
            e["date"] for e in timesheet_entries
            if e["job_type"].lower() == "leave"
        }

        for leave_date, leave_type_id in leave_date_map.items():
            if leave_date not in existing_dates:
                timesheet_entries.append({
                    "date": leave_date,
                    "job_type": "Leave",
                    "leave_type_id": leave_type_id,
                    "work_hours": 0,
                    "ot_hours": 0,
                    "break_hours": 0,
                    "description": "Approved Leave",
                    "is_auto_leave": True,  # 🔥 important for UI
                })

        med_certs = MedicalCertificate.objects.filter(
            employee=current_employee,
            month=current_month
        )

        leaves_locked = True
        # Prepare context
        context = {
            "current_month": current_month,
            "task_types": task_types,
            "tasks": tasks,
            "leave_types": leave_types,
            "task_types_json": mark_safe(json.dumps(list(task_types.values("id", "name")))),
            "tasks_json": mark_safe(json.dumps(list(tasks.values("id", "name", "type_id", "description")))),
            "leave_types_json": mark_safe(json.dumps(list(leave_types.values("id", "name")))),
            "timesheet_data_json": mark_safe(json.dumps(timesheet_entries)),
            "read_only": json.dumps(read_only),
            "leaves_locked": json.dumps(leaves_locked),
            "is_pmo": json.dumps(current_employee.is_pmo),
            "ot_allowed": json.dumps(current_employee.ot_allowed),
            "public_holidays": holiday_map,
            "ph_allowed": json.dumps(current_employee.ph_allowed),
            "medical_files": med_certs,
            "current_employee": current_employee,
            "submission_status_text":submission_status_text,
            "user": request.user,
            "existing_remark" : submission.employee_remark if submission else "",
        }

        return render(request, "employees/timesheet.html", context)

    # ---------------------------
    # POST → Save or Submit timesheet
    # ---------------------------
    # --------------- POST Method ---------------
    elif request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            month_str = data.get("month")
            is_submit = data.get("submit", False)
            remark = data.get("remark")

            # --- Month range ---
            first_day = datetime.datetime.strptime(f"{month_str}-01", "%Y-%m-%d").date()
            last_day = (first_day + relativedelta(months=1)) - datetime.timedelta(days=1)

            # --- Period record ---
            period, _ = TimesheetPeriod.objects.get_or_create(
                period_type="MONTH",
                start_date=first_day,
                end_date=last_day
            )

            # --- Submission record ---
            submission, _ = TimesheetSubmission.objects.get_or_create(
                employee=current_employee,
                period=period,
                defaults={"status": "DRAFT"}
            )

            if submission.status in ["APPROVED"]:
                return JsonResponse({
                    "status": "error",
                    "message": f"Timesheet already {submission.status.lower()}. Editing not allowed."
                }, status=403)

            from django.db.models import Q
            valid_conditions = []

            approved_leaves = LeaveApplication.objects.filter(
                employee=current_employee,
                status="APPROVED",
                start_date__lte=last_day,
                end_date__gte=first_day,
            )
            leave_date_map = {}  # date -> leave_type_id
            for leave in approved_leaves:
                current = max(leave.start_date, first_day)
                end = min(leave.end_date, last_day)
                for work_date in iter_working_dates(current, end):
                    leave_date_map[work_date] = leave.leave_type_id

            entries = data.get("entries", [])
            def safe_int(val):
                try:
                    return int(val)
                except (TypeError, ValueError):
                    return None

            for entry in entries:
                entry_date = parse_date(entry["date"])
                job_type = (entry.get("job_type") or "Work").strip()
                task_id = safe_int(entry.get("task_id"))
                task_type_id = safe_int(entry.get("task_type_id"))
                leave_type_id = safe_int(entry.get("leave_type_id"))
                work_hours = float(entry.get("work_hours") or 0)
                ot_hours = float(entry.get("ot_hours") or 0)
                break_hours = float(entry.get("break_hours") or 0)
                desc = entry.get("description", "")
                is_ot = str(entry.get("ot_flag", "No")).lower() in ("yes", "true", "1")

                # === 1️⃣ Work entries (OT + non-OT) ===
                if job_type.lower() == "work":
                    obj, created = TimesheetEntry.objects.update_or_create(
                        employee=current_employee,
                        job_type = "Work",
                        date=entry_date,
                        task_id=task_id,
                        is_ot=is_ot,
                        defaults={
                            "task_type_id": task_type_id,
                            "job_type": "Work",
                            "work_hours": work_hours,
                            "ot_hours": ot_hours,
                            "break_hours": break_hours,
                            "leave_type_id": leave_type_id,
                            "description": desc,
                            "timesheet_period": period,
                        },
                    )
                    valid_conditions.append(Q(date=entry_date, task_id=task_id, is_ot=is_ot))
                    continue

                # === 2️⃣ Leave entries (partial-day supported) ===
                if job_type.lower() == "leave":
                    # Leave entries are controlled via Leave Applications, not editable here.
                    continue
                is_public_holiday = PublicHoliday.objects.filter(holiday_date=entry_date, is_active=True).exists()
                if is_public_holiday and not current_employee.ph_work_allowed:
                    # force entry to be Public Holiday with zero hours
                    job_type = "Public Holiday"
                    work_hours = 0
                    ot_hours = 0
                    leave_type_id = None
                # === 3️⃣ Public holidays ===
                if job_type.lower() == "public holiday":
                    obj, created = TimesheetEntry.objects.update_or_create(
                        employee=current_employee,
                        date=entry_date,
                        job_type="Public Holiday",
                        defaults={
                            "task_id": None,
                            "task_type_id": None,
                            "is_ot": False,
                            "work_hours": work_hours,
                            "ot_hours": ot_hours,
                            "break_hours": break_hours,
                            "leave_type_id": None,
                            "description": desc or "Public Holiday",
                            "timesheet_period": period,
                        },
                    )
                    valid_conditions.append(Q(date=entry_date, job_type="Public Holiday"))
                    continue

                # === 4️⃣ Fallback (untyped work with no task) ===
                obj, created = TimesheetEntry.objects.update_or_create(
                    employee=current_employee,
                    date=entry_date,
                    task_id=None,
                    is_ot=is_ot,
                    defaults={
                        "task_type_id": None,
                        "job_type": "Work",
                        "work_hours": work_hours,
                        "ot_hours": ot_hours,
                        "break_hours": break_hours,
                        "leave_type_id": leave_type_id,
                        "description": desc or "General Work",
                        "timesheet_period": period,
                    },
                )
                valid_conditions.append(Q(date=entry_date, task_id__isnull=True, is_ot=is_ot))

            # === Sync approved leaves into timesheet entries (source of truth) ===
            for leave_date, leave_type_id in leave_date_map.items():
                TimesheetEntry.objects.update_or_create(
                    employee=current_employee,
                    date=leave_date,
                    job_type="Leave",
                    leave_type_id=leave_type_id,
                    defaults={
                        "task_id": None,
                        "task_type_id": None,
                        "is_ot": False,
                        "work_hours": 0,
                        "ot_hours": 0,
                        "break_hours": 0,
                        "description": "Approved Leave",
                        "timesheet_period": period,
                    },
                )
                valid_conditions.append(Q(date=leave_date, job_type="Leave", leave_type_id=leave_type_id))

            # === Cleanup of stale entries ===
            base_qs = TimesheetEntry.objects.filter(
                employee=current_employee,
                date__range=(first_day, last_day),
            )
            if valid_conditions:
                from functools import reduce
                combined_q = reduce(lambda a, b: a | b, valid_conditions)
                base_qs.exclude(combined_q).delete()
            else:
                base_qs.delete()

            submission.employee_remark = remark
            submission.save(update_fields=["employee_remark"])

            # === Finalize submission ===
            if is_submit:
                # MEDICAL_LEAVE_IDS = list(
                #     LeaveType.objects.filter(name__icontains="Medical").values_list("id", flat=True))
                # entries = TimesheetEntry.objects.filter(
                #     employee=current_employee,
                #     date__range=(first_day, last_day),
                #     job_type = "Leave",
                #     leave_type_id__in=MEDICAL_LEAVE_IDS,
                # )
                #
                # has_mc_uploaded = MedicalCertificate.objects.filter(
                #     employee=current_employee,
                #     month=month_str,
                # ).exists()
                # if entries and not has_mc_uploaded:
                #     return JsonResponse({
                #         "status": "error",
                #         "message": "You have selected Medical Leave, but no Medical Certificate is uploaded."
                #     })

                leaves = TimesheetEntry.objects.filter(
                    employee=current_employee,
                    date__range=(first_day, last_day),
                    job_type="Leave")

                entries  = TimesheetEntry.objects.filter(
                    employee=current_employee,
                    date__range=(first_day, last_day),
                    job_type="Work")


                if len(entries) < 1 and len(leaves) == 0:
                    return JsonResponse({
                        "status": "error",
                        "message": "You must enter hours OR report leaves to submit the timesheet."
                    })


                update_leave_cycle(current_employee)
                submission.status = "SUBMITTED"
                submission.submitted_at = timezone.now()
                submission.save()

                TimesheetApprovalLog.objects.create(
                    submission=submission,
                    actor=current_employee.user,
                    action="SUBMIT",
                    comment="Timesheet submitted for manager review.",
                )
                msg = "✅ Timesheet submitted successfully."
            else:
                submission.status = "DRAFT"
                submission.save()

                TimesheetApprovalLog.objects.create(
                    submission=submission,
                    actor=current_employee.user,
                    action="SAVE_DRAFT",
                    comment="Timesheet saved as draft.",
                )
                msg = "💾 Timesheet saved as draft."

            return JsonResponse({
                "status": "success",
                "redirect": reverse("saved_timesheets"),
                "message": f"Timesheet {'submitted' if is_submit else 'saved'} successfully."
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"status": "error", "message": str(e)}, status=500)





# employees/views.py
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from django.conf import settings
import os

def export_timesheet_file(request, submission_id):
    """
    Export filled timesheet for the given submission ID
    (for now: static data only)
    """
    # Normally you'll fetch employee/submission details from DB:
    # submission = TimesheetSubmission.objects.get(id=submission_id)
    # emp = submission.employee
    # For now, we'll use static placeholders.

    template_path = os.path.join(settings.BASE_DIR, 'YM_Timesheet_Template.xlsx')
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Time Sheet"]

    # --- Static header values ---
    static_values = {
        "C7": "E001 - Rajesh Kumar",
        "C9": "DXC Technology",
        "C11": "October 2025",
        "C13": "15-Mar-2023",
        "C15": datetime.date.today().strftime("%d-%b-%Y"),
        "N7": "John Smith",
        "N9": "S1234567A",
        "N11": "+65 9123 4567",
        "N13": "2 years 7 months",
    }
    for cell_ref, val in static_values.items():
        ws[cell_ref] = val

    # --- Weekly blocks ---
    block_starts = ["A", "G", "M", "S", "Y"]
    start_row = 19
    week_starts = [
        datetime.date(2025, 9, 29),
        datetime.date(2025, 10, 6),
        datetime.date(2025, 10, 13),
        datetime.date(2025, 10, 20),
        datetime.date(2025, 10, 27),
    ]
    static_wh = 8
    static_bh = 1
    static_ot = 0.5

    for i, start_col in enumerate(block_starts):
        if i >= len(week_starts):
            break
        start_idx = column_index_from_string(start_col)
        cols = [get_column_letter(start_idx + offset) for offset in range(5)]
        week_start = week_starts[i]
        for r in range(7):
            d = week_start + datetime.timedelta(days=r)
            ws[f"{cols[0]}{start_row + r}"] = d.strftime("%d-%b-%Y")
            ws[f"{cols[1]}{start_row + r}"] = d.strftime("%a")
            ws[f"{cols[2]}{start_row + r}"] = static_wh
            ws[f"{cols[3]}{start_row + r}"] = static_bh
            ws[f"{cols[4]}{start_row + r}"] = static_ot

    # Prepare response (no need to save locally)
    response = HttpResponse(
        content_type='application/vnd.ms-excel.sheet.macroEnabled.12'
    )
    response['Content-Disposition'] = f'attachment; filename="Timesheet_{submission_id}.xlsx"'
    wb.save(response)
    return response


from django.shortcuts import render
from .models import TimesheetSubmission, MainClient, EndClient, Employee
from django.db.models import Q

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_timesheets(request):
    """
    Admin view — displays all Submitted and Approved timesheets
    with filters for Main Client, End Client, Employee Name, and Month.
    """

    # Get filter params from GET request
    main_client_id = request.GET.get("main_client")
    end_client_id = request.GET.get("end_client")
    employee_name = request.GET.get("employee_name")
    month = request.GET.get("month")
    selected_year = request.GET.get("year", "")

    # Base queryset — only submitted or approved
    submissions = TimesheetSubmission.objects.filter(
        status__in=["SUBMITTED", "APPROVED", "REJECTED"]
    ).select_related("employee", "period", "employee__main_account", "employee__end_client")

    # Apply filters
    if main_client_id:
        submissions = submissions.filter(employee__main_account_id=main_client_id)
    if end_client_id:
        submissions = submissions.filter(employee__end_client_id=end_client_id)
    if employee_name:
        submissions = submissions.filter(employee__full_name__icontains=employee_name)
    if selected_year:
        submissions = submissions.filter(period__start_date__year=selected_year)
    if month:
        submissions = submissions.filter(period__start_date__month=month)

    # Fetch dropdown data
    main_clients = MainClient.objects.all()
    end_clients = EndClient.objects.all()

    months = [
        ("01", "January"), ("02", "February"), ("03", "March"), ("04", "April"),
        ("05", "May"), ("06", "June"), ("07", "July"), ("08", "August"),
        ("09", "September"), ("10", "October"), ("11", "November"), ("12", "December"),
    ]

    current_year = datetime.datetime.now().year
    years = [str(y) for y in range(2025, current_year + 2)]

    selected_month = request.GET.get("month", "")
    selected_year = request.GET.get("year", "")

    context = {
        "submissions": submissions,
        "main_clients": main_clients,
        "end_clients": end_clients,
        "months": months,
        "years": years,
        "parent_template":"employees/dashboard.html",
        "filters": {
            "main_client": main_client_id,
            "end_client": end_client_id,
            "employee_name": employee_name or "",
            "month": selected_month,
            "year": selected_year,
        },
    }

    return render(request, "employees/admin_timesheets.html", context)


from django.http import HttpResponse
from .timesheet_exporter import TimesheetExporter
import tempfile

def export_timesheet(request, employee_id, month, year):
    """
    Exports a single employee's timesheet as an Excel file.
    """
    exporter = TimesheetExporter(employee_id, month, year)

    # Save to temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        exporter.export(tmp.name)
        tmp.seek(0)
        response = HttpResponse(
            tmp.read(),
            content_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )
        filename = f"Timesheet_{exporter.employee.full_name}_{month}_{year}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response



import zipfile
import io
import tempfile
from django.http import HttpResponse
from django.shortcuts import redirect

import io
import zipfile
import tempfile
import datetime
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
import calendar
from employees.business_logic.aws_service import send_ses_email_with_attachment
from .timesheet_exporter import TimesheetExporter


def export_multiple_timesheets(request):
    if request.method != "POST":
        return redirect("admin_timesheets")

    # -------------------------
    # Incoming POST values
    # -------------------------
    selected = request.POST.getlist("selected_timesheets")
    send_to = request.POST.get("send_to_email")
    excel_email_export = request.POST.get("email_export")        # Excel email
    pdf_export = request.POST.get("pdf_export")                 # PDF download
    pdf_email_export = request.POST.get("email_pdf_export")     # PDF email
    mc_export = request.POST.get("mc_export")

    if not selected:
        messages.error(request, "No timesheets selected.")
        return redirect("admin_timesheets")

    # -----------------------------------------------------------
    # SINGLE EXCEL DOWNLOAD REDIRECT (only when no export flags)
    # -----------------------------------------------------------

    if (
            len(selected) == 1
            and not pdf_export
            and not pdf_email_export
            and not excel_email_export
            and not mc_export  # ✅ IMPORTANT
    ):

        emp_id, mon, yr = selected[0].split("|")
        return redirect("export_timesheet", employee_id=emp_id, month=mon, year=yr)

    # ===========================================================
    # CASE MC EXPORT (Download ZIP only)
    # ===========================================================
    if mc_export == "1":
        mc_zip_buffer = io.BytesIO()
        found_any = False

        with zipfile.ZipFile(mc_zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for sel in selected:
                emp_id, mon, yr = sel.split("|")
                emp = Employee.objects.get(id=int(emp_id))

                # get MC files for this employee and month
                month_str = f"{yr}-{int(mon):02d}"
                mc_files = MedicalCertificate.objects.filter(
                    employee=emp,
                    month=month_str
                )

                if mc_files.exists():
                    found_any = True

                safe_name = ''.join(c for c in emp.full_name[:10] if c.isalnum())
                counter = 1

                for mc in mc_files:
                    if mc.file:
                        ext = os.path.splitext(mc.file.name)[1]
                        file_name = f"{safe_name}_{counter}{ext}"

                        try:
                            with mc.file.open("rb") as f:
                                zipf.writestr(file_name, f.read())
                        except Exception as e:
                            print("MC ERROR:", e)

                        counter += 1

        if not found_any:
            messages.error(request, "No medical certificates found for selected employees.")
            return redirect("admin_timesheets")

        mc_zip_buffer.seek(0)
        final_mc_zip = mc_zip_buffer.getvalue()

        response = HttpResponse(final_mc_zip, content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="MedicalCertificates_YMG.zip"'
        return response

    # ===========================================================
    # CASE A: PDF Export (single or multiple) - Download or Email
    # ===========================================================


    if pdf_export == "1" or pdf_email_export == "1":
        # table_rows summary used in email body
        table_rows = []

        # ----- SINGLE SELECTED: return single PDF or email single PDF -----
        if len(selected) == 1:
            sel = selected[0]
            emp_id, mon, yr = sel.split("|")
            emp_id, mon, yr = int(emp_id), int(mon), int(yr)

            # Add summary row for email
            try:
                emp_obj = Employee.objects.get(id=emp_id)
                month_name = calendar.month_name[mon]
                table_rows.append({
                    "name": emp_obj.full_name,
                    "month_year": f"{month_name} {yr}",
                })
            except Employee.DoesNotExist:
                table_rows = []

            exporter = TimesheetExporter(emp_id, mon, yr)

            tmp_excel_path = None
            tmp_pdf_path = None
            try:
                # Create temp Excel
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                    tmp_excel_path = tmp_excel.name
                exporter.export(tmp_excel_path)

                # PDF path
                tmp_pdf_path = tmp_excel_path.replace(".xlsx", ".pdf")

                # Convert to PDF (uses your exporter method and LibreOffice)
                exporter.export_excel_and_pdf(tmp_excel_path, tmp_pdf_path)

                # Read PDF bytes
                with open(tmp_pdf_path, "rb") as f:
                    pdf_bytes = f.read()

                filename = f"{exporter.employee.full_name}_{yr}_{mon:02d}.pdf"

                # EMAIL single PDF
                if pdf_email_export == "1":
                    if not send_to:
                        messages.error(request, "Please enter an email address.")
                        return redirect("admin_timesheets")

                    try:
                        send_ses_email_with_attachment(
                            recipient=send_to,
                            filename=filename,
                            file_bytes=pdf_bytes,
                            table_rows=table_rows
                        )
                        messages.success(request, f"PDF exported & emailed to {send_to} successfully!")
                    except Exception as e:
                        messages.error(request, f"Email failed: {str(e)}")

                    return redirect("admin_timesheets")

                # DOWNLOAD single PDF
                response = HttpResponse(pdf_bytes, content_type="application/pdf")
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response

            finally:
                # clean up temp files if they exist
                try:
                    if tmp_excel_path and os.path.exists(tmp_excel_path):
                        os.remove(tmp_excel_path)
                except Exception:
                    pass
                try:
                    if tmp_pdf_path and os.path.exists(tmp_pdf_path):
                        os.remove(tmp_pdf_path)
                except Exception:
                    pass

        # ----- MULTIPLE SELECTED: build ZIP of PDFs -----
        pdf_zip_buffer = io.BytesIO()

        with zipfile.ZipFile(pdf_zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for sel in selected:
                emp_id, mon, yr = sel.split("|")
                emp_id, mon, yr = int(emp_id), int(mon), int(yr)

                # populate table_rows with readable summary for email
                try:
                    emp_obj = Employee.objects.get(id=emp_id)
                    month_name = calendar.month_name[mon]
                    table_rows.append({
                        "name": emp_obj.full_name,
                        "month_year": f"{month_name} {yr}",
                    })
                except Employee.DoesNotExist:
                    # skip adding to table_rows but continue export
                    pass

                exporter = TimesheetExporter(emp_id, mon, yr)

                tmp_excel_path = None
                tmp_pdf_path = None
                try:
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_excel:
                        tmp_excel_path = tmp_excel.name
                    exporter.export(tmp_excel_path)

                    tmp_pdf_path = tmp_excel_path.replace(".xlsx", ".pdf")
                    exporter.export_excel_and_pdf(tmp_excel_path, tmp_pdf_path)

                    with open(tmp_pdf_path, "rb") as f:
                        filename = f"{exporter.employee.full_name}_{yr}_{mon:02d}.pdf"
                        zipf.writestr(filename, f.read())
                finally:
                    # remove temporary files
                    try:
                        if tmp_excel_path and os.path.exists(tmp_excel_path):
                            os.remove(tmp_excel_path)
                    except Exception:
                        pass
                    try:
                        if tmp_pdf_path and os.path.exists(tmp_pdf_path):
                            os.remove(tmp_pdf_path)
                    except Exception:
                        pass

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"YMGTimesheets_PDF_{timestamp}.zip"
        pdf_zip_buffer.seek(0)
        final_pdf_zip = pdf_zip_buffer.getvalue()

        # EMAIL multiple PDFs as ZIP
        if pdf_email_export == "1":
            if not send_to:
                messages.error(request, "Please enter an email address.")
                return redirect("admin_timesheets")

            try:
                send_ses_email_with_attachment(
                    recipient=send_to,
                    filename=zip_filename,
                    file_bytes=final_pdf_zip,
                    table_rows=table_rows
                )
                messages.success(request, f"PDF export emailed to {send_to} successfully!")
            except Exception as e:
                messages.error(request, f"Email failed: {str(e)}")

            return redirect("admin_timesheets")

        # DOWNLOAD ZIP of PDFs
        response = HttpResponse(final_pdf_zip, content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'
        return response

    # ===========================================================
    # CASE B: EXCEL ZIP Export (Download or Email)
    # ===========================================================
    excel_zip_buffer = io.BytesIO()
    table_rows = []  # reset / build for excel email

    with zipfile.ZipFile(excel_zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for sel in selected:
            emp_id, mon, yr = sel.split("|")
            emp = Employee.objects.get(id=int(emp_id))

            # populate summary row for email
            month_name = calendar.month_name[int(mon)]
            table_rows.append({
                "name": emp.full_name,
                "month_year": f"{month_name} {yr}",
            })

            emp_id, mon, yr = int(emp_id), int(mon), int(yr)

            exporter = TimesheetExporter(emp_id, mon, yr)

            tmp_excel_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp_excel_path = tmp.name
                exporter.export(tmp_excel_path)

                with open(tmp_excel_path, "rb") as f:
                    filename = f"{exporter.employee.full_name}_{yr}_{mon:02d}.xlsx"
                    zipf.writestr(filename, f.read())
            finally:
                try:
                    if tmp_excel_path and os.path.exists(tmp_excel_path):
                        os.remove(tmp_excel_path)
                except Exception:
                    pass

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_zip_filename = f"YMGTimesheets_{timestamp}.zip"
    excel_zip_buffer.seek(0)
    final_excel_zip = excel_zip_buffer.getvalue()

    # EMAIL Excel ZIP
    if excel_email_export == "1":
        if not send_to:
            messages.error(request, "Please enter an email address.")
            return redirect("admin_timesheets")

        try:
            send_ses_email_with_attachment(
                recipient=send_to,
                filename=excel_zip_filename,
                file_bytes=final_excel_zip,
                table_rows=table_rows
            )
            messages.success(request, f"Export emailed to {send_to} successfully!")
        except Exception as e:
            messages.error(request, f"Email failed: {str(e)}")

        return redirect("admin_timesheets")


    # DOWNLOAD Excel ZIP
    response = HttpResponse(final_excel_zip, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{excel_zip_filename}"'
    return response


@login_required
def upload_medical_certificate(request):
    employee = get_object_or_404(Employee, email=request.user.username)

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST required"}, status=400)

    file = request.FILES.get("file")
    if file.size > 1 * 1024 * 1024:  # 1 MB
        return JsonResponse({"error": "File too large"}, status=400)

    if file.content_type not in ["image/jpeg", "image/png", "application/pdf"]:
        return JsonResponse({"error": "Invalid file type"}, status=400)
    month = request.POST.get("month")

    if not file:
        return JsonResponse({"status": "error", "message": "No file uploaded"}, status=400)

    # Find period (if already created)
    first_day = datetime.datetime.strptime(f"{month}-01", "%Y-%m-%d").date()
    period = TimesheetPeriod.objects.filter(start_date=first_day).first()

    mc = MedicalCertificate.objects.create(
        employee=employee,
        period=period,
        month=month,
        file=file
    )

    return JsonResponse({
        "status": "success",
        "file_id": mc.id,
        "file_name": mc.file.name.split("/")[-1],
        "file_url": settings.MEDIA_URL + mc.file.name
    })


@login_required
def list_medical_certificates(request):
    emp_id = request.GET.get("employee_id", "")
    month = request.GET.get("month")
    if emp_id and request.user.is_staff:
        employee = get_object_or_404(Employee, id=emp_id)
    else:
        # normal employee
        employee = get_object_or_404(Employee, email=request.user.username)
    files = MedicalCertificate.objects.filter(employee=employee, month=month)

    return JsonResponse([
        {"id": f.id, "name": f.file.name.split("/")[-1], "url": f.file.url}
        for f in files
    ], safe=False)


@login_required
def delete_medical_certificate(request):
    employee = get_object_or_404(Employee, email=request.user.username)

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST required"}, status=400)

    file_id = request.POST.get("id")

    MedicalCertificate.objects.filter(id=file_id, employee=employee).delete()

    return JsonResponse({"status": "success"})

from .business_logic.leave_service import *
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseForbidden
from django.db.models import Sum
from datetime import date

from employees.models import (
    Employee,
    LeaveApplication,
    LeaveAdjustment,
    LeaveType,
)
from employees.business_logic.leave_service import get_calendar_cycle

from django.shortcuts import get_object_or_404, render
from django.http import HttpResponseForbidden
from django.db.models import Sum

from employees.models import (
    Employee,
    LeaveApplication,
    LeaveAdjustment,
)
from employees.business_logic.leave_service import get_calendar_cycle

from decimal import Decimal
def employee_leave_details(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    # ===========================
    # ACCESS CONTROL
    # ===========================
    if request.user.is_staff or request.user.is_superuser:
        template = "employees/employee_leave_details.html"
        parent_template = "employees/dashboard.html"
        role = "admin"

    elif hasattr(request.user, "manager_profile"):
        manager = request.user.manager_profile
        if employee.client_account_manager != manager:
            return HttpResponseForbidden("Not allowed")

        template = "employees/employee_leave_details.html"
        parent_template = "employees/manager_dashboard.html"
        role = "manager"

    elif hasattr(request.user, "employee_profile"):
        if request.user != employee.user:
            return HttpResponseForbidden("Not allowed")

        template = "employees/employee_leave_details_basic.html"
        parent_template = "employees/base.html"
        role = "employee"

    else:
        return HttpResponseForbidden("Not allowed")

    # ===========================
    # CALENDAR YEAR
    # ===========================
    cycle = get_calendar_cycle()
    cycle_start = cycle["start"]
    cycle_end = cycle["end"]

    # ===========================
    # ENTITLEMENTS (FROM EMPLOYEE)
    # ===========================
    annual_entitled = Decimal(employee.yearly_annual_leave_entitlement or 0)
    medical_entitled = Decimal(employee.yearly_sick_leave_entitlement or 0)

    # ===========================
    # APPROVED LEAVES (SOURCE OF TRUTH)
    # ===========================
    annual_used = Decimal("0.0")
    medical_used = Decimal("0.0")

    approved_leaves = LeaveApplication.objects.filter(
        employee=employee,
        status="APPROVED",
        start_date__lte=cycle_end,
        end_date__gte=cycle_start
    ).select_related("leave_type")

    for leave in approved_leaves:
        start = max(leave.start_date, cycle_start)
        end = min(leave.end_date, cycle_end)

        leave_name = leave.leave_type.name.lower()
        days = Decimal(
            str(
                effective_leave_days(
                    start,
                    end,
                    is_half_day=("half" in leave_name),
                )
            )
        )

        if "annual" in leave_name:
            annual_used += days
        elif "medical" in leave_name:
            medical_used += days

    # ===========================
    # MANUAL ADJUSTMENTS
    # ===========================
    adjustments = LeaveAdjustment.objects.filter(
        employee=employee,
        created_at__date__gte=cycle_start,
        created_at__date__lte=cycle_end
    ).order_by("-created_at")

    annual_adjustment_total = adjustments.filter(
        adjustment_type="ANNUAL"
    ).aggregate(total=Sum("days"))["total"] or 0

    medical_adjustment_total = adjustments.filter(
        adjustment_type="MEDICAL"
    ).aggregate(total=Sum("days"))["total"] or 0

    # ===========================
    # FINAL BALANCES
    # ===========================
    annual_balance = annual_entitled - annual_used + annual_adjustment_total
    medical_balance = medical_entitled - medical_used + medical_adjustment_total

    # ===========================
    # LEAVE HISTORY (APPLICATIONS ONLY)
    # ===========================
    leave_history = []

    leave_apps = LeaveApplication.objects.filter(
        employee=employee
    ).order_by("-applied_on")

    for leave in leave_apps:
        days = effective_leave_days(
            leave.start_date,
            leave.end_date,
            is_half_day=("half" in leave.leave_type.name.lower()),
        )

        leave_history.append({
            "from": leave.start_date,
            "to": leave.end_date,
            "type": leave.leave_type.name,
            "days": days,
            "status": leave.status,
        })

    # ===========================
    # RECENT LEAVES
    # ===========================
    recent_leaves = leave_apps[:3]

    # ===========================
    # RENDER
    # ===========================
    return render(request, template, {
        "employee": employee,

        # SUMMARY
        "annual_entitled": annual_entitled,
        "annual_used": round(annual_used, 1),
        "annual_adjustment_total": annual_adjustment_total,
        "annual_balance": round(annual_balance, 1),

        "medical_entitled": medical_entitled,
        "medical_used": round(medical_used, 1),
        "medical_adjustment_total": medical_adjustment_total,
        "medical_balance": round(medical_balance, 1),

        # DETAILS
        "adjustments": adjustments,
        "leave_history": leave_history,
        "recent_leaves": recent_leaves,

        # UI FLAGS
        "parent_template": parent_template,
        "is_admin": role == "admin",
        "is_manager": role == "manager",
        "is_employee_self": role == "employee",
    })


@login_required
def employee_timesheets(request, pk):
    employee = get_object_or_404(Employee, pk=pk)

    # ============================
    # ADMIN ACCESS
    # ============================
    if request.user.is_staff or request.user.is_superuser:
        parent_template = "employees/dashboard.html"

    # ============================
    # MANAGER ACCESS
    # ============================
    elif hasattr(request.user, "manager_profile"):
        manager = request.user.manager_profile

        # Ensure manager only accesses own employees
        if employee.client_account_manager != manager:
            return HttpResponseForbidden("Not allowed")

        parent_template = "employees/manager_dashboard.html"

    # ============================
    # EMPLOYEES NOT ALLOWED
    # ============================
    else:
        return HttpResponseForbidden("Not allowed")

    # ============================
    # FETCH TIMESHEETS
    # ============================
    submissions = TimesheetSubmission.objects.filter(
        employee=employee,
        status__in=["SUBMITTED", "APPROVED", "REJECTED"]
    ).select_related("period").order_by("-period__start_date")

    return render(request, "employees/employee_timesheets.html", {
        "employee": employee,
        "submissions": submissions,
        "parent_template": parent_template,
    })


def employee_claims(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, "employees/employee_claims.html", {
        "employee": employee
    })


@login_required
def leave_adjustment_add(request, pk):
    if not request.user.is_staff:
        return redirect("employee_leave_details", pk=pk)

    employee = get_object_or_404(Employee, pk=pk)
    cycle = get_or_create_active_cycle(employee)

    if request.method == "POST":
        form = LeaveAdjustmentForm(request.POST)
        if form.is_valid():
            adj = form.save(commit=False)
            adj.employee = employee
            adj.cycle = cycle
            adj.approved_by = request.user
            adj.save()

            # Recalculate the cycle immediately
            update_leave_cycle(employee)

            return redirect("employee_leave_details", pk=pk)
    else:
        form = LeaveAdjustmentForm()

    return render(request, "employees/leave_adjustment_add.html", {
        "form": form,
        "employee": employee,
    })


from employees.models import LeaveCycle, PublicHoliday
# Employee Homepage
@login_required
def employee_home(request):
    employee = get_object_or_404(Employee, email=request.user.username)

    # Get active leave cycle
    leave_cycle = LeaveCycle.objects.filter(employee=employee).order_by("-cycle_start").first()
    country = employee.location or "Singapore"
    # Upcoming PH (next 3 months)
    upcoming_holidays = PublicHoliday.objects.filter(
        country=country,
        holiday_date__gte=datetime.date.today(),
        holiday_date__lte=datetime.date.today() + datetime.timedelta(days=90)
    ).order_by("holiday_date")

    return render(request, "employees/employee_home.html", {
        "employee": employee,
        "leave_cycle": leave_cycle,
        "upcoming_holidays": upcoming_holidays,
    })


@login_required
def claim_submit(request):
    employee = get_object_or_404(Employee, email=request.user.username)

    if request.method == "POST":
        form = ClaimForm(request.POST)
        attachment_form = ClaimAttachmentForm(request.POST, request.FILES)

        if form.is_valid():
            claim = form.save(commit=False)
            claim.employee = employee
            claim.status = "SUBMITTED"
            claim.save()

            # Save multiple attachments
            files = request.FILES.getlist("file")
            for f in files:
                if f.size > 1 * 1024 * 1024:  # 1 MB
                    return JsonResponse({"error": "File too large"}, status=400)

                if f.content_type not in ["image/jpeg", "image/png", "application/pdf"]:
                    return JsonResponse({"error": "Invalid file type"}, status=400)

                ClaimAttachment.objects.create(claim=claim, file=f)

            return redirect("claim_list")  # we'll build this next

    else:
        form = ClaimForm()
        attachment_form = ClaimAttachmentForm()

    return render(request, "employees/claim_submit.html", {
        "form": form,
        "attachment_form": attachment_form,
        "employee": employee,
    })


@login_required
def claim_list(request):
    employee = get_object_or_404(Employee, email=request.user.username)

    claims = employee.claims.all().order_by("-submitted_at")

    return render(request, "employees/claim_list.html", {
        "claims": claims,
        "employee": employee,
    })


@login_required
def claim_detail(request, claim_id):
    claim = get_object_or_404(Claim, id=claim_id)

    # Ensure employee can only view their own claim
    employee = get_object_or_404(Employee, email=request.user.username)
    if claim.employee != employee:
        return HttpResponseForbidden("You cannot view this claim.")

    attachments = claim.attachments.all()

    return render(request, "employees/claim_detail.html", {
        "claim": claim,
        "attachments": attachments,
    })

@login_required
def admin_claims(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("Not allowed")

    name_query = request.GET.get("name", "").strip()
    month_param = request.GET.get("month", "").strip()
    employee_param = request.GET.get("employee", "").strip()
    employee_param_name = ""
    if employee_param:
        try:
            employee_param_name= Employee.objects.get(id = employee_param).full_name
        except Exception as e:
            employee_param_name = ""


    claims = Claim.objects.select_related("employee").order_by("-submitted_at")

    if employee_param:
        claims = claims.filter(employee_id=employee_param)
    # Filter by employee name
    if name_query:
        claims = claims.filter(employee__full_name__icontains=name_query)

    # Filter by month (YYYY-MM)
    if month_param and "-" in month_param:
        try:
            year, month = month_param.split("-")
            claims = claims.filter(start_date__year=year, start_date__month=month)
        except:
            pass

    return render(request, "employees/admin_claims.html", {
        "claims": claims,
        "name_query": name_query,
        "month": month_param,
        "employee_param": employee_param,
        "employee_param_name": employee_param_name,
    })


@login_required
def admin_claim_detail(request, claim_id):
    if not request.user.is_staff:
        return HttpResponseForbidden("Not allowed")

    claim = get_object_or_404(Claim.objects.select_related("employee"), id=claim_id)
    attachments = ClaimAttachment.objects.filter(claim=claim)
    employee_filter = request.GET.get("employee", "")
    return render(request, "employees/claim_detail.html", {
        "claim": claim,
        "attachments": attachments,
        "is_admin": True,
        "employee_filter": employee_filter, # 👈 important
    })


@login_required
def admin_reject_claim(request, claim_id):
    if not request.user.is_staff:
        return HttpResponseForbidden("Not allowed")

    claim = get_object_or_404(Claim, id=claim_id)

    if request.method == "POST":
        comment = request.POST.get("rejected_comment", "").strip()

        claim.status = "REJECTED"
        claim.rejected_comment = comment
        claim.rejected_at = timezone.now()
        claim.save()

        return redirect("admin_claim_detail", claim_id=claim.id)

    return render(request, "employees/admin_reject_claim.html", {
        "claim": claim
    })

from django.shortcuts import render
from django.db.models import Count, Sum
from django.utils import timezone
from datetime import date, timedelta
import json

from employees.models import (
    Employee, TimesheetEntry, TimesheetSubmission, Claim, LeaveRecord, MainClient, EndClient, RoleSow
)
from django.contrib.auth.decorators import user_passes_test
from django.utils.timezone import localdate
from dateutil.relativedelta import relativedelta


def active_user_dashboard(request):
    return render(request, "employees/active_dashboard.html")

def can_review(request, submission):
    # Admin allowed
    if request.user.is_staff or request.user.is_superuser:
        return True

    # Manager allowed only if this employee belongs to them
    if hasattr(request.user, "manager_profile"):
        return submission.employee.client_account_manager == request.user.manager_profile

    return False

@login_required
def approve_timesheet(request, submission_id):
    submission = get_object_or_404(TimesheetSubmission, id=submission_id)

    if not can_review(request, submission):
        return HttpResponseForbidden("Not allowed")

    # Determine manager (if applicable)
    manager = request.user.manager_profile if hasattr(request.user, "manager_profile") else None

    submission.mark_approved(manager, comment="Approved.")
    TimesheetApprovalLog.objects.create(
        submission=submission,
        actor=request.user,
        action="APPROVE",
        comment="Approved."
    )

    messages.success(request, f"{submission.employee.full_name}'s timesheet APPROVED.")

    # Redirect correctly depending on role
    if request.user.is_staff:
        return redirect("admin_timesheets")
    else:
        return redirect("manager_timesheets")

@login_required
def reject_timesheet(request, submission_id):
    submission = get_object_or_404(TimesheetSubmission, id=submission_id)

    if not can_review(request, submission):
        return HttpResponseForbidden("Not allowed")

    comment = request.POST.get("reason", "Timesheet rejected.")

    manager = request.user.manager_profile if hasattr(request.user, "manager_profile") else None

    submission.mark_rejected(manager, comment)
    TimesheetApprovalLog.objects.create(
        submission=submission,
        actor=request.user,
        action="REJECT",
        comment=comment
    )

    messages.error(request, f"{submission.employee.full_name}'s timesheet REJECTED.")

    # Redirect to manager or admin view
    if request.user.is_staff:
        return redirect("admin_timesheets")
    else:
        return redirect("manager_timesheets")


def public_holidays(request):
    selected_country = request.GET.get("country", "Singapore")

    current_year = datetime.date.today().year
    next_year = current_year + 1
    previous_year = current_year - 1

    holidays_current = PublicHoliday.objects.filter(
        country=selected_country,
        holiday_date__year=current_year,
        is_active=True
    ).order_by("holiday_date")

    holidays_next = PublicHoliday.objects.filter(
        country=selected_country,
        holiday_date__year=next_year,
        is_active=True
    ).order_by("holiday_date")

    holidays_previous = PublicHoliday.objects.filter(
        country=selected_country,
        holiday_date__year=previous_year,
        is_active=True
    ).order_by("holiday_date")


    countries = [c[0] for c in PublicHoliday.COUNTRY_CHOICES]
    # Clean up duplicates caused by spacing / casing
    countries = sorted({c.strip().title() for c in countries})
    return render(request, "employees/public_holidays.html", {
        "selected_country": selected_country,
        "holidays_current": holidays_current,
        "holidays_next": holidays_next,
        "holidays_previous": holidays_previous,
        "countries": countries,
        "current_year": current_year,
        "next_year": next_year,
        "previous_year": previous_year,
    })


def add_public_holiday(request):
    if request.method == "POST":
        form = PublicHolidayForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.country = obj.country.strip().title()
            obj.save()
            messages.success(request, "Holiday added successfully!")
            return redirect("public_holidays")
    else:
        form = PublicHolidayForm(initial={"country": "Singapore"})

    return render(request, "employees/public_holiday_form.html", {
        "form": form,
        "mode": "add"
    })


def edit_public_holiday(request, pk):
    holiday = PublicHoliday.objects.get(pk=pk)

    if request.method == "POST":
        form = PublicHolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.country = obj.country.strip().title()
            obj.save()
            messages.success(request, "Holiday updated successfully!")
            return redirect("public_holidays")
    else:
        form = PublicHolidayForm(instance=holiday)

    return render(request, "employees/public_holiday_form.html", {
        "form": form,
        "mode": "edit",
        "holiday": holiday
    })


def delete_public_holiday(request, pk):
    holiday = PublicHoliday.objects.get(pk=pk)
    holiday.delete()
    messages.success(request, "Holiday deleted.")
    return redirect("public_holidays")


@login_required
def manager_employee_list(request):
    manager = request.user.manager_profile

    employees = Employee.objects.filter(
        client_account_manager=manager,
        is_active=True
    ).order_by('full_name')

    return render(request, "employees/manager_employee_list.html", {
        "employees": employees
    })

@login_required
def manager_timesheets(request):
    manager = request.user.manager_profile

    # Employees under this manager
    team_ids = Employee.objects.filter(
        client_account_manager=manager
    ).values_list("id", flat=True)

    # -----------------------------
    # Get filter parameters
    # -----------------------------
    main_client_id = request.GET.get("main_client")
    end_client_id = request.GET.get("end_client")
    employee_name = request.GET.get("employee_name")
    month = request.GET.get("month")
    selected_year = request.GET.get("year", "")

    # -----------------------------
    # Manager-scoped queryset
    # -----------------------------
    submissions = TimesheetSubmission.objects.filter(
        employee_id__in=team_ids,
        status__in=["SUBMITTED", "APPROVED", "REJECTED"]
    ).select_related("employee", "period", "employee__main_account", "employee__end_client")

    # -----------------------------
    # Apply filters (same as admin)
    # -----------------------------
    if main_client_id:
        submissions = submissions.filter(employee__main_account_id=main_client_id)

    if end_client_id:
        submissions = submissions.filter(employee__end_client_id=end_client_id)

    if employee_name:
        submissions = submissions.filter(employee__full_name__icontains=employee_name)

    if selected_year:
        submissions = submissions.filter(period__start_date__year=selected_year)

    if month:
        submissions = submissions.filter(period__start_date__month=month)

    # -----------------------------
    # DROPDOWN OPTIONS (LIMITED)
    # -----------------------------

    # Manager has only ONE main client
    main_clients = MainClient.objects.filter(id=manager.main_client_id)

    # Manager can have multiple end clients
    end_clients = manager.end_clients.all()

    # -----------------------------
    # Month & Year dropdowns (same as admin)
    # -----------------------------
    months = [
        ("01", "January"), ("02", "February"), ("03", "March"), ("04", "April"),
        ("05", "May"), ("06", "June"), ("07", "July"), ("08", "August"),
        ("09", "September"), ("10", "October"), ("11", "November"), ("12", "December"),
    ]

    current_year = datetime.datetime.now().year
    years = [str(y) for y in range(2025, current_year + 2)]

    # -----------------------------
    # Render template
    # -----------------------------
    return render(request, "employees/admin_timesheets.html", {
        "submissions": submissions,
        "main_clients": main_clients,        # LIMITED
        "end_clients": end_clients,          # LIMITED
        "months": months,
        "years": years,
        "is_manager": True,
        "parent_template": "employees/manager_dashboard.html",

        "filters": {
            "main_client": main_client_id,
            "end_client": end_client_id,
            "employee_name": employee_name or "",
            "month": month or "",
            "year": selected_year or "",
        },
    })




@login_required
def manager_leave_details(request):
    manager = request.user.manager_profile
    team_ids = Employee.objects.filter(client_account_manager=manager).values_list("id", flat=True)

    leaves = LeaveRecord.objects.filter(employee_id__in=team_ids).select_related("employee", "leave_type")

    return render(request, "employees/manager_leave_details.html", {
        "leaves": leaves
    })



@login_required
def leave_history(request):
    user = request.user

    # =====================================================
    # DETERMINE ROLE + EMPLOYEE ACCESS
    # =====================================================
    if user.is_staff or user.is_superuser:
        role = "admin"
        parent_template = "employees/dashboard.html"
        employees_qs = Employee.objects.all()

    elif hasattr(user, "manager_profile"):
        role = "manager"
        parent_template = "employees/manager_dashboard.html"

        manager = user.manager_profile

        # Manager can ONLY see employees under them
        employees_qs = Employee.objects.filter(client_account_manager=manager)

    else:
        return HttpResponseForbidden("Not allowed")

    # =====================================================
    # FILTER INPUTS
    # =====================================================
    employee_name = request.GET.get("employee_name", "").strip()
    month = request.GET.get("month", "").strip()
    leave_type_filter = request.GET.get("leave_type", "").strip()

    # Last 3 months
    today = timezone.now().date()
    start_date = today - datetime.timedelta(days=90)

    # =====================================================
    # BASE QUERIES
    # =====================================================
    ts_entries = TimesheetEntry.objects.filter(
        employee__in=employees_qs,
        job_type="Leave",
        date__gte=start_date
    ).select_related("employee", "leave_type")

    imported_entries = LeaveImport.objects.filter(
        employee__in=employees_qs,
        date__gte=start_date
    ).select_related("employee", "leave_type")

    # =====================================================
    # APPLY FILTERS
    # =====================================================
    if employee_name:
        ts_entries = ts_entries.filter(employee__full_name__icontains=employee_name)
        imported_entries = imported_entries.filter(employee__full_name__icontains=employee_name)

    if month:
        ts_entries = ts_entries.filter(date__month=month)
        imported_entries = imported_entries.filter(date__month=month)

    if leave_type_filter:
        ts_entries = ts_entries.filter(leave_type_id=leave_type_filter)
        imported_entries = imported_entries.filter(leave_type_id=leave_type_filter)

    # =====================================================
    # MERGE & NORMALIZE OUTPUT
    # =====================================================
    # Generate last 3 months for dropdown
    months = []
    cursor = today.replace(day=1)

    for _ in range(3):
        month_num = cursor.strftime("%m")
        month_name = cursor.strftime("%B")
        months.append((month_num, month_name))
        cursor = (cursor - datetime.timedelta(days=1)).replace(day=1)

    combined = []

    for t in ts_entries:
        combined.append({
            "employee": t.employee.full_name,
            "employee_id": t.employee.id if t.employee else None,
            "date": t.date,
            "type": t.leave_type.name if t.leave_type else "-",
            "source": "Timesheet",
        })

    for r in imported_entries:
        combined.append({
            "employee": r.employee.full_name,
            "employee_id": r.employee.id if r.employee else None,
            "date": r.date,
            "type": r.leave_type.name if r.leave_type else "-",
            "source": "Imported",
        })

    combined.sort(key=lambda x: x["date"], reverse=True)

    # =====================================================
    # RENDER
    # =====================================================
    return render(request, "employees/leave_history.html", {
        "entries": combined,
        "parent_template": parent_template,
        "role": role,

        # filters
        "employee_name": employee_name,
        "month": month,
        "leave_type_filter": leave_type_filter,
        "months": months,
        # dropdown
        "leave_types": LeaveType.objects.all(),
    })


def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email")

        try:
            user = User.objects.get(username=email)
        except User.DoesNotExist:
            return render(request, "employees/forgot_password.html", {
                "email_sent": True  # security: don't reveal existence
            })

        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        reset_url = request.build_absolute_uri(
            reverse("reset_password", kwargs={
                "uidb64": uid,
                "token": token
            })
        )

        # 🔹 SEND EMAIL USING SES
        send_password_reset_email(
            recipient=email,
            reset_url=reset_url
        )

        return render(request, "employees/forgot_password.html", {
            "email_sent": True
        })

    return render(request, "employees/forgot_password.html")


import re

def is_strong_password(password):
    if len(password) < 8:
        return "Password must be at least 8 characters long."
    if not re.search(r"[A-Z]", password):
        return "Password must contain at least one uppercase letter."
    if not re.search(r"[a-z]", password):
        return "Password must contain at least one lowercase letter."
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        return "Password must contain at least one special character."
    return None



def reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except Exception:
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        messages.error(request, "Invalid or expired reset link.")
        return redirect("login")

    if request.method == "POST":
        password1 = request.POST.get("password")
        password2 = request.POST.get("confirm_password")

        if not password1 or not password2:
            messages.error(request, "Password cannot be empty.")
            return render(request, "employees/reset_password.html")

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return render(request, "employees/reset_password.html")

        strength_error = is_strong_password(password1)
        if strength_error:
            messages.error(request, strength_error)
            return render(request, "employees/reset_password.html")

        user.set_password(password1)
        user.save(update_fields=["password"])

        messages.success(request, "Password reset successful. Please login.")
        return redirect("login")

    return render(request, "employees/reset_password.html")


@login_required
def employee_configure_tasks(request):
    # Must be an employee
    if not hasattr(request.user, "employee_profile"):
        return HttpResponseForbidden("Only employees can access this page")

    employee = request.user.employee_profile

    # Employee MUST have a RoleSOW
    if not employee.role_sow:
        return HttpResponseForbidden("No Role/SOW assigned to you")

    rolesow = employee.role_sow

    # 🔹 Task Types (same as admin)
    task_types = rolesow.task_types.all()

    tasks = Task.objects.filter(
        role_sow=rolesow, is_active=True
    )

    return render(
        request,
        "employees/configurables_role_details.html",
        {
            "rolesow": rolesow,
            "task_types": task_types,
            "tasks": tasks,
            "is_employee_view": True,   # 🔑 key flag
        }
    )

def employee_add_task(request):
    employee = request.user.employee_profile
    rolesow = employee.role_sow

    if request.method == "POST":
        form = TaskForm(request.POST, rolesow=rolesow)
        if form.is_valid():
            task = form.save(commit=False)
            task.role_sow = rolesow
            task.associated_employee = employee
            task.is_base = False
            task.save()
            return redirect("employee_configure_tasks")
    else:
        form = TaskForm(rolesow=rolesow)


    return render(
        request,
        "employees/add_task.html",
        {
            "form": form,
            "rolesow": rolesow,
            "is_employee_view": True,  # 🔑 REQUIRED
        }
    )



@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def toggle_task_active(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    # Toggle only is_active
    task.is_active = not task.is_active
    task.save(update_fields=["is_active"])

    return redirect(
        "configurables_role_details",
        rolesow_id=task.role_sow.id
    )

from employees.models import AssetInventory, Asset
@login_required
def employee_assets(request):
    if not hasattr(request.user, "employee_profile"):
        return HttpResponseForbidden("Only employees allowed")

    employee = request.user.employee_profile

    assets = Asset.objects.select_related("inventory").filter(
        employee=employee
    )

    return render(
        request,
        "employees/employee_assets.html",
        {
            "assets": assets
        }
    )


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_assets(request):
    assets = Asset.objects.select_related(
        "inventory", "employee"
    ).all()

    return render(
        request,
        "employees/admin_assets.html",
        {
            "assets": assets
        }
    )

from django.db.models import Prefetch
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_inventory(request):
    inventory = AssetInventory.objects.prefetch_related(
        Prefetch(
            "assignments",
            queryset=Asset.objects.filter(is_with_employee=True),
            to_attr="active_assignments"
        )
    )

    return render(
        request,
        "employees/admin_inventory.html",
        {
            "inventory": inventory
        }
    )

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def assign_asset(request):
    if request.method == "POST":
        Asset.objects.create(
            inventory_id=request.POST["inventory"],
            employee_id=request.POST["employee"],
            received_on=request.POST["received_on"],
            is_with_employee=True
        )
        return redirect("admin_assets")

    return render(
        request,
        "employees/assign_asset.html",
        {
            "inventory": AssetInventory.objects.filter(is_active=True),
            "employees": Employee.objects.filter(is_active=True),
            "selected_inventory": request.GET.get("inventory"),
        }
    )


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def add_inventory(request):
    if request.method == "POST":
        AssetInventory.objects.create(
            asset_type=request.POST["asset_type"],
            asset_color=request.POST.get("asset_color"),
            serial_number=request.POST.get("serial_number"),
            purchased_at=request.POST.get("purchased_at"),
            purchased_date=request.POST.get("purchased_date") or None,
            optional_description=request.POST.get("optional_description"),
            last_repaired_on=request.POST.get("last_repaired_on") or None,
            last_repair_details=request.POST.get("last_repair_details"),
            is_active="is_active" in request.POST,
        )
        return redirect("admin_inventory")

    return render(
        request,
        "employees/add_inventory.html",
        {
            "is_edit": False
        }
    )


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def edit_inventory(request, inventory_id):
    inventory_item = get_object_or_404(AssetInventory, id=inventory_id)

    if request.method == "POST":
        inventory_item.asset_type = request.POST["asset_type"]
        inventory_item.asset_color = request.POST.get("asset_color")
        inventory_item.serial_number = request.POST.get("serial_number")
        inventory_item.purchased_at = request.POST.get("purchased_at")
        inventory_item.purchased_date = request.POST.get("purchased_date") or None
        inventory_item.optional_description = request.POST.get("optional_description")
        inventory_item.last_repaired_on = request.POST.get("last_repaired_on") or None
        inventory_item.last_repair_details = request.POST.get("last_repair_details")
        inventory_item.is_active = "is_active" in request.POST
        inventory_item.save()

        return redirect("admin_inventory")

    return render(
        request,
        "employees/add_inventory.html",
        {
            "inventory_item": inventory_item,
            "is_edit": True
        }
    )


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def edit_assignment(request, asset_id):
    assignment = get_object_or_404(Asset, id=asset_id)

    if request.method == "POST":
        assignment.returned_on = request.POST.get("returned_on") or None
        assignment.is_with_employee = "is_with_employee" in request.POST
        assignment.remarks = request.POST.get("remarks")
        assignment.save()

        return redirect("admin_inventory")

    return render(
        request,
        "employees/edit_assignment.html",
        {
            "assignment": assignment
        }
    )


from .forms import LeaveApplicationForm
from .models import LeaveApplication


@login_required
def apply_leave(request):
    employee = request.user.employee_profile

    if request.method == "POST":
        form = LeaveApplicationForm(request.POST, employee=employee)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.duration = effective_leave_days(
                leave.start_date,
                leave.end_date,
                is_half_day=("half" in leave.leave_type.name.lower()),
            )
            leave.save()

            return redirect(
                "employee_leave_details",
                employee.id
            )

    else:
        form = LeaveApplicationForm(employee=employee)

    return render(request, "employees/leave_apply.html", {
        "form": form,
        "existing_leaves": LeaveApplication.objects.filter(employee=employee)
            .exclude(status="REJECTED")
            .select_related("leave_type")
            .order_by("-start_date")[:20],
    })
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

@staff_member_required
def admin_leave_approvals(request):
    """
    Shows only PENDING leave applications for admin action.
    """
    leaves = LeaveApplication.objects.filter(
        status="PENDING"
    ).select_related("employee", "leave_type").order_by("applied_on")
    recent_actions = LeaveApplication.objects.filter(
        approved_on__date=timezone.now().date()
    ).exclude(status="PENDING").order_by("-approved_on")[:5]
    return render(
        request,
        "employees/admin_leave_approvals.html",
        {
            "leaves": leaves,
            "recent_actions": recent_actions
        }
    )


@staff_member_required
def approve_leave(request, leave_id):
    leave = get_object_or_404(LeaveApplication, id=leave_id)

    leave.status = "APPROVED"
    leave.approved_by = request.user
    leave.approved_on = timezone.now()
    leave.save()

    return redirect("admin_leave_approvals")


@staff_member_required
def reject_leave(request, leave_id):
    leave = get_object_or_404(LeaveApplication, id=leave_id)

    if request.method == "POST":
        comment = request.POST.get("comment", "").strip()

        if not comment:
            return redirect("admin_leave_approvals")

        leave.status = "REJECTED"
        leave.admin_comment = comment
        leave.approved_by = request.user
        leave.approved_on = timezone.now()
        leave.save()

        return redirect("admin_leave_approvals")

    return redirect("admin_leave_approvals")
