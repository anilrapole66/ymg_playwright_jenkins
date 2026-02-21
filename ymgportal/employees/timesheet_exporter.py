import os
import subprocess
import sys
from datetime import date, timedelta
from django.db.models import Sum

# ✅ Set up Django before importing any models
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "employee_portal.settings")

import django
django.setup()

from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from django.conf import settings
from employees.models import Employee, TimesheetEntry, TimesheetPeriod, LeaveAdjustment
import os
from employees.business_logic.leave_service import update_leave_cycle

class TimesheetExporter:
    """
    Clean OOP version of dynamic timesheet export.
    Builds date matrix (Mon–Sun weeks) and fills data into Excel template.
    """

    def __init__(self, employee_id, month, year=None):
        self.employee_id = employee_id
        self.month = int(month)
        self.year = year or date.today().year
        self.employee = Employee.objects.select_related(
            "main_account", "end_client", "client_account_manager"
        ).get(id=employee_id)

        # Determine TimesheetPeriod (if exists)
        self.period = TimesheetPeriod.objects.filter(
            start_date__month=self.month, start_date__year=self.year
        ).first()

        # Load all entries for this employee/month
        self.entries = TimesheetEntry.objects.filter(
            employee=self.employee,
            date__month=self.month,
            date__year=self.year,
        )
        # Build {date: entry_data_dict}
        self.entry_map = self._prepare_entry_map()

        # Compute all dates (Mon→Sun blocks)
        self.all_dates = self._generate_date_blocks()

    # ----------------------------------------------------------------
    def _calculate_service_period(self, joining_date):
        """Return period of service as Years, Months, Days."""
        today = date.today()

        # Days difference
        delta = today - joining_date
        days = delta.days

        # Compute Y/M/D manually
        years = days // 365
        remaining_days = days % 365
        months = remaining_days // 30
        days = remaining_days % 30

        return years, months, days

    def _prepare_entry_map(self):
        """Combine multiple entries for the same date:
        - Sum work_hours and ot_hours
        - Override break_hours (last non-zero value wins)
        """
        entry_map = {}

        for entry in self.entries:
            d = entry.date
            if d not in entry_map:
                entry_map[d] = {
                    "work_hours": 0.0,
                    "break_hours": 0.0,
                    "ot_hours": 0.0,
                    "leave": "",
                }

            # Sum work and OT hours
            entry_map[d]["work_hours"] += float(entry.work_hours or 0)
            entry_map[d]["ot_hours"] += float(entry.ot_hours or 0)

            # Override break hours (if provided)
            if entry.break_hours not in (None, 0, 0.0):
                entry_map[d]["break_hours"] = float(entry.break_hours)

            # If any entry for that date is a leave, mark it
            if entry.job_type == "Leave":
                if entry.leave_type.name == "Annual Leave":
                    entry_map[d]["leave"] = "AL"
                if entry.leave_type.name == "Medical Leave":
                    entry_map[d]["leave"] = "SL"
                if entry.leave_type.name == "Half Annual Leave":
                    entry_map[d]["leave"] = "HAL"
                if entry.leave_type.name == "Half Medical Leave":
                    entry_map[d]["leave"] = "HSL"
                if entry.leave_type.name == "No Pay Leave":
                    entry_map[d]["leave"] = "NPL"
                if entry.leave_type.name == "Other Leave":
                    entry_map[d]["leave"] = "OL"


        return entry_map

    # ----------------------------------------------------------------
    def _generate_date_blocks(self):
        """Build all Mon–Sun dates covering the full month."""
        from calendar import monthrange

        first_day = date(self.year, self.month, 1)
        last_day = date(self.year, self.month, monthrange(self.year, self.month)[1])

        # Move back to previous Monday
        start_monday = first_day - timedelta(days=(first_day.weekday()))
        # Move forward to next Sunday
        end_sunday = last_day + timedelta(days=(6 - last_day.weekday()))

        current = start_monday
        all_days = []
        while current <= end_sunday:
            all_days.append(current)
            current += timedelta(days=1)
        return all_days

    # ----------------------------------------------------------------
    def _fill_excel_headers(self, ws):
        """Fills static header section"""
        emp = self.employee
        period_label = self.period.start_date.strftime("%B %Y") if self.period else f"{self.month}/{self.year}"

        ws["C7"] = f"{emp.full_name}"
        ws["C9"] = emp.main_account.name if emp.main_account else "-"
        ws["C11"] = period_label
        ws["C13"] = emp.date_of_joining.strftime("%d-%b-%Y") if emp.date_of_joining else "-"
        ws["C15"] = date.today().strftime("%d-%b-%Y")

        ws["N7"] = "-"
        ws["N9"] = "-"
        ws["N11"] = emp.phone or "-"
        if emp.date_of_joining:
            y, m, d = self._calculate_service_period(emp.date_of_joining)
            ws["N13"] = f"{y} years {m} months {d} days"
        else:
            ws["N13"] = "-"
            

    # ----------------------------------------------------------------
    def _fill_excel_weeks(self, ws):
        """Fills all date blocks dynamically."""
        block_starts = ["A", "G", "M", "S", "Y"]
        start_row = 19

        for block_index, week_start_index in enumerate(range(0, len(self.all_dates), 7)):
            if block_index >= len(block_starts):
                break

            start_col = block_starts[block_index]
            start_idx = column_index_from_string(start_col)
            cols = [get_column_letter(start_idx + offset) for offset in range(6)]

            for day_offset in range(7):
                if week_start_index + day_offset >= len(self.all_dates):
                    break
                d = self.all_dates[week_start_index + day_offset]
                entry = self.entry_map.get(d, {"work_hours": 0, "break_hours": 0, "ot_hours": 0, "leave": ""})

                row = start_row + day_offset
                ws[f"{cols[0]}{row}"] = d.strftime("%d-%b-%Y")
                ws[f"{cols[1]}{row}"] = d.strftime("%a")

                ws[f"{cols[2]}{row}"] = entry["work_hours"]
                ws[f"{cols[3]}{row}"] = entry["break_hours"]
                ws[f"{cols[4]}{row}"] = entry["ot_hours"]
                ws[f"{cols[5]}{row}"] = entry["leave"]

    def _fill_leave_summary(self, ws):
        """Fills Earned / Taken / Balance for Annual & Medical leave."""

        cycle = update_leave_cycle(self.employee)

        annual_adjustments = float(
            LeaveAdjustment.objects.filter(
                employee=self.employee,
                cycle=cycle,
                adjustment_type="ANNUAL"
            ).aggregate(total=Sum("days"))["total"] or 0
        )

        # C column → Earned
        ws["C30"] = cycle.annual_leave_earned
        ws["C32"] = cycle.medical_leave_earned

        # I column → Used / Taken
        ws["I30"] = cycle.annual_leave_used
        ws["I32"] = cycle.medical_leave_used

        # O column → Balance
        ws["O30"] = cycle.annual_leave_balance

        if annual_adjustments != 0:
            formatted_adj = f"{annual_adjustments:+}"  # → +2 or -1
            ws["O30"] = f"{cycle.annual_leave_balance} (Adjusted: {formatted_adj})"

        ws["O32"] = cycle.medical_leave_balance

    def export_excel_and_pdf(self, output_excel_path, output_pdf_path):
        """Save Excel, then convert to PDF using LibreOffice."""
        # 1) Save Excel
        wb = self.export(output_excel_path)

        # 2) Convert to PDF
        subprocess.run([
            "libreoffice",
            "--headless",
            "--convert-to",
            "pdf:calc_pdf_Export:{" \
            "ReduceImageResolution:false," \
            "SelectPdfVersion:1," \
            "Quality:100," \
            "UseLosslessCompression:true," \
            "ExportFormFields:true," \
            "Scale:175," \
            "}",
            "--outdir", os.path.dirname(output_pdf_path),
            output_excel_path
        ])


    # ----------------------------------------------------------------
    def export(self, output_path=None):
        """Generate the Excel workbook and optionally save to disk."""
        template_path = os.path.join(settings.BASE_DIR, 'YM_Timesheet_Template.xlsx')
        wb = load_workbook(template_path, keep_vba=True)
        ws = wb["Time Sheet"]

        self._fill_excel_headers(ws)
        self._fill_excel_weeks(ws)
        self._fill_leave_summary(ws)
        if output_path:
            wb.save(output_path)
        return wb


# ----------------------------------------------------------------
# ✅ Test block (run this file directly)
# ----------------------------------------------------------------
if __name__ == "__main__":

    # Change to real values in your DB
    EMPLOYEE_ID = 411
    MONTH = 11
    YEAR = 2025

    exporter = TimesheetExporter(EMPLOYEE_ID, MONTH, YEAR)
    print("📅 Date range:", exporter.all_dates[0], "→", exporter.all_dates[-1])
    print("📊 Entries found:", len(exporter.entries))
    print("🗓 Generating workbook...")

    output_file = os.path.join(settings.BASE_DIR, f"Timesheet_Test_{MONTH}_{YEAR}.xlsm")
    pdf_file = os.path.join(settings.BASE_DIR, f"Timesheet_Test_{MONTH}_{YEAR}.pdf")
    exporter.export(output_file)
    exporter.export_excel_and_pdf(output_file, pdf_file)
