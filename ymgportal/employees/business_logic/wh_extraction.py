import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# === CONFIGURATION ===
input_folder = r"/home/jai/Documents/Timesheets"  # 👈 Change this path
output_file = "Consolidated_Timesheet_WH_Only.xlsx"

# Column groups (Date, WH)
COLUMN_BLOCKS = [
    ("A", "C"),
    ("G", "I"),
    ("M", "O"),
    ("S", "U"),
    ("Y", "AA"),
]

def extract_timesheet_data(file_path):
    wb = load_workbook(file_path, data_only=True)
    if "Time Sheet" not in wb.sheetnames:
        print(f"❌ Skipping {file_path} (No 'Time Sheet' sheet found)")
        return None

    ws = wb["Time Sheet"]

    # --- Header Info ---
    name = ws["C7"].value
    month_text = str(ws["C11"].value).strip()
    doj = ws["C13"].value

    # Extract target month (e.g., "Oct 2025" → month=10, year=2025)
    try:
        target_date = datetime.strptime(month_text, "%b %Y")
    except ValueError:
        target_date = datetime.strptime(month_text, "%B %Y")
    target_month = target_date.month
    target_year = target_date.year

    # --- Leave Info ---
    al_earned = ws["C30"].value
    sl_earned = ws["C32"].value
    al_taken = ws["I30"].value
    sl_taken = ws["I32"].value
    al_balance = ws["O30"].value
    sl_balance = ws["O32"].value

    # --- Daily WH data ---
    daily_hours = {}
    for row in range(22, 55):  # Includes all rows with date blocks
        for (date_col, wh_col) in COLUMN_BLOCKS:
            date_val = ws[f"{date_col}{row}"].value
            wh_val = ws[f"{wh_col}{row}"].value

            if not date_val:
                continue

            # Parse Excel date
            date_obj = None
            if isinstance(date_val, datetime):
                date_obj = date_val
            else:
                try:
                    date_obj = datetime.strptime(str(date_val), "%d-%b-%y")
                except Exception:
                    continue

            # --- FILTER: Skip if not same month ---
            if date_obj.month != target_month or date_obj.year != target_year:
                continue

            date_str = date_obj.strftime("%d-%b")
            daily_hours[date_str] = wh_val or 0

    # --- Combine all data ---
    data = {
        "Employee Name": name,
        "Month": month_text,
        "Date of Joining": doj,
        "AL_Earned": al_earned,
        "SL_Earned": sl_earned,
        "AL_Taken": al_taken,
        "SL_Taken": sl_taken,
        "AL_Balance": al_balance,
        "SL_Balance": sl_balance,
    }
    data.update(daily_hours)
    return data


# === MAIN LOOP ===
records = []
for file in os.listdir(input_folder):
    if file.lower().endswith((".xlsm", ".xlsx")):
        full_path = os.path.join(input_folder, file)
        print(f"Processing: {file}")
        record = extract_timesheet_data(full_path)
        if record:
            records.append(record)

# === COMBINE & SAVE ===
if records:
    df = pd.DataFrame(records)

    # Sort columns (1-Oct, 2-Oct, etc.)
    static_cols = ["Employee Name", "Month", "Date of Joining",
                   "AL_Earned", "SL_Earned", "AL_Taken",
                   "SL_Taken", "AL_Balance", "SL_Balance"]

    day_cols = sorted([c for c in df.columns if "-" in c],
                      key=lambda x: int(x.split("-")[0]))
    df = df[static_cols + day_cols]

    df.to_excel(output_file, index=False)
    print(f"✅ Clean WH-only data saved to: {output_file}")
else:
    print("⚠️ No valid timesheets found.")
