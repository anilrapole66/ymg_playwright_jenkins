import os
import pandas as pd
from openpyxl import load_workbook

# === CONFIGURATION ===
input_folder = r"/home/jai/Documents/Timesheets"  # 👈 Change this path
output_file = "Consolidated_Timesheet.xlsx"


# Column groups (Date, WH, BH, OTH, LD)
COLUMN_BLOCKS = [
    ("A", "C", "D", "E", "F"),
    ("G", "I", "J", "K", "L"),
    ("M", "O", "P", "Q", "R"),
    ("S", "U", "V", "W", "X"),
    ("Y", "AA", "AB", "AC", "AD"),
]


def extract_timesheet_data(file_path):
    wb = load_workbook(file_path, data_only=True)
    if "Time Sheet" not in wb.sheetnames:
        print(f"❌ Skipping {file_path} (No 'Time Sheet' sheet found)")
        return None

    ws = wb["Time Sheet"]

    # --- Header Info ---
    name = ws["C7"].value
    month = ws["C11"].value
    doj = ws["C13"].value

    # --- Leave Info ---
    al_earned = ws["C30"].value
    sl_earned = ws["C32"].value
    al_taken = ws["I30"].value
    sl_taken = ws["I32"].value
    al_balance = ws["O30"].value
    sl_balance = ws["O32"].value

    # --- Daily WH/BH/OTH/LD data ---
    daily_data = {}
    for row in range(22, 50):  # Adjust if your sheet extends beyond row 50
        for (date_col, wh_col, bh_col, oth_col, ld_col) in COLUMN_BLOCKS:
            date_val = ws[f"{date_col}{row}"].value
            if not date_val:
                continue

            # Format the date (e.g. 02-Oct)
            try:
                date_str = date_val.strftime("%d-%b")
            except Exception:
                date_str = str(date_val).strip()

            # Extract WH/BH/OTH/LD
            wh = ws[f"{wh_col}{row}"].value or 0
            bh = ws[f"{bh_col}{row}"].value or 0
            oth = ws[f"{oth_col}{row}"].value or 0
            ld = ws[f"{ld_col}{row}"].value or ""

            # Store in dict as separate columns
            daily_data[f"{date_str}_WH"] = wh
            daily_data[f"{date_str}_BH"] = bh
            daily_data[f"{date_str}_OTH"] = oth
            daily_data[f"{date_str}_LD"] = ld

    # --- Combine all data into one dict ---
    data = {
        "Employee Name": name,
        "Month": month,
        "Date of Joining": doj,
        "AL_Earned": al_earned,
        "SL_Earned": sl_earned,
        "AL_Taken": al_taken,
        "SL_Taken": sl_taken,
        "AL_Balance": al_balance,
        "SL_Balance": sl_balance,
    }
    data.update(daily_data)
    return data


# === MAIN LOOP ===
records = []
for file in os.listdir(input_folder):
    if file.lower().endswith((".xlsm", ".xlsx")):
        full_path = os.path.join(input_folder, file)
        print(f"Processing: {file}")
        record = extract_timesheet_data(full_path)
        if record:
            records.append(record)

# === COMBINE & SAVE ===
if records:
    df = pd.DataFrame(records)

    # Order columns logically
    static_cols = ["Employee Name", "Month", "Date of Joining",
                   "AL_Earned", "SL_Earned", "AL_Taken",
                   "SL_Taken", "AL_Balance", "SL_Balance"]

    # Sort day columns numerically (1-Oct to 31-Oct)
    day_cols = sorted([c for c in df.columns if "-" in c],
                      key=lambda x: int(x.split("-")[0]))
    df = df[static_cols + day_cols]

    df.to_excel(output_file, index=False)
    print(f"✅ Consolidated data saved to: {output_file}")
else:
    print("⚠️ No valid timesheets found.")
